from google.cloud import bigquery
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

PROJECT  = "staff-sizing-portfolio"
DATASET  = "staff_sizing"
OUTPUT   = "sheets/hiring_reco_2027.xlsx"

client = bigquery.Client(project=PROJECT)

# ── 1. Exportar datos desde BigQuery ──────────────────────────────────────
print("Exportando datos desde BigQuery...")

df_detail = client.query(f"""
  SELECT
    base_code            AS Base,
    role                 AS Rol,
    category             AS Categoria,
    ROUND(current_fte,1) AS FTE_Actual,
    min_headcount        AS Dotacion_Minima,
    optimal_headcount    AS Dotacion_Optima,
    ROUND(worst_gap_base,1) AS Peor_Gap,
    months_with_gap      AS Meses_Con_Gap,
    hires_needed_base    AS Contrataciones_Base,
    hires_needed_stress  AS Contrataciones_Estres,
    ROUND(avg_training_days,0) AS Dias_Capacitacion,
    ROUND(months_to_fill,1)    AS Meses_Para_Cubrir,
    hiring_priority      AS Prioridad,
    priority_rank        AS Ranking
  FROM `{PROJECT}.{DATASET}.mart_hiring_reco_2027`
  ORDER BY priority_rank ASC
""").to_dataframe()

print(f"  {len(df_detail)} filas exportadas")

# ── 2. Crear Excel con pandas (3 hojas base) ──────────────────────────────
print("Creando archivo Excel...")

os.makedirs("sheets", exist_ok=True)

with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:

    # ── Pestaña DETALLE ──────────────────────────────────────────────────
    df_detail.to_excel(writer, sheet_name="DETALLE", index=False)

    # ── Pestaña RESUMEN ──────────────────────────────────────────────────
    p1_count = len(df_detail[df_detail["Prioridad"].str.contains("P1", na=False)])
    p2_count = len(df_detail[df_detail["Prioridad"].str.contains("P2", na=False)])
    p3_count = len(df_detail[df_detail["Prioridad"].str.contains("P3", na=False)])
    p4_count = len(df_detail[df_detail["Prioridad"].str.contains("P4", na=False)])

    df_summary = pd.DataFrame({
        "Metrica": [
            "Total contrataciones necesarias (escenario base)",
            "Total contrataciones necesarias (escenario estres)",
            "Roles con prioridad P1 URGENTE",
            "Roles con prioridad P2 CRITICO",
            "Roles con prioridad P3 PLANIFICAR",
            "Roles con prioridad P4 OK",
            "Promedio dias de capacitacion",
            "Base con mayor brecha",
            "Rol mas critico",
        ],
        "Valor": [
            int(df_detail["Contrataciones_Base"].sum()),
            int(df_detail["Contrataciones_Estres"].sum()),
            p1_count,
            p2_count,
            p3_count,
            p4_count,
            round(df_detail["Dias_Capacitacion"].mean(), 1),
            df_detail.iloc[0]["Base"],
            df_detail.iloc[0]["Rol"],
        ]
    })
    df_summary.to_excel(writer, sheet_name="RESUMEN", index=False)

    # Top 10 críticos en RESUMEN
    df_top10 = df_detail[
        df_detail["Prioridad"].str.contains("P1|P2", na=False)
    ].head(10)[[
        "Base", "Rol", "Categoria", "Contrataciones_Base",
        "Contrataciones_Estres", "Prioridad", "Meses_Para_Cubrir"
    ]]
    df_top10.to_excel(writer, sheet_name="RESUMEN", index=False, startrow=14)

    # ── Pestaña SUPUESTOS ────────────────────────────────────────────────
    df_supuestos = pd.DataFrame({
        "Variable": [
            "Nivel de dotacion objetivo",
            "Escenario base",
            "Escenario estres",
            "Formula FTE proyectado",
            "Formula GAP",
            "Lead time default",
            "Nivel de servicio FTE",
            "Fuente de datos",
            "Ultima actualizacion",
            "",
            "PRIORIDADES DE CONTRATACION",
            "P1 URGENTE",
            "P2 CRITICO",
            "P3 PLANIFICAR",
            "P4 OK",
        ],
        "Descripcion": [
            "Dotacion minima requerida por base y rol",
            "Rotacion historica promedio 2022-2024",
            "Rotacion alta en temporada (promedio + 1.5 stddev)",
            "FTE_actual - Salidas_mes - (Ausentismo_semanal x 4)",
            "FTE_proyectado - min_headcount_requerido",
            "14 dias cuando no hay dato del proveedor",
            "Dotacion minima operacional",
            "BigQuery - staff_sizing.mart_hiring_reco_2027",
            pd.Timestamp.now().strftime("%Y-%m-%d"),
            "",
            "",
            "Gap < -10 Y 10+ meses con gap -> iniciar proceso inmediato",
            "Gap < -5 O 6+ meses con gap -> proximo ciclo de seleccion",
            "Gap < 0 -> planificar Q3/Q4 2027",
            "Sin gap proyectado -> sin accion inmediata",
        ]
    })
    df_supuestos.to_excel(writer, sheet_name="SUPUESTOS", index=False)

print(f"  Archivo creado: {OUTPUT}")

# ── 3. Aplicar formato con openpyxl ───────────────────────────────────────
print("Aplicando formato...")

wb = load_workbook(OUTPUT)

COLOR_HEADER = "1F3864"
COLOR_P1     = "FF4444"
COLOR_P2     = "FF9900"
COLOR_P3     = "FFFF00"
COLOR_P4     = "92D050"
COLOR_WHITE  = "FFFFFF"
COLOR_GRAY   = "F2F2F2"

def style_header_row(ws, row=1, color=COLOR_HEADER):
    for cell in ws[row]:
        cell.font      = Font(bold=True, color=COLOR_WHITE, size=11)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value else 0 for cell in col),
            default=0
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(
            min_width, min(max_len + 2, max_width)
        )

def add_border(cell):
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Formato pestaña DETALLE ───────────────────────────────────────────────
ws_detail = wb["DETALLE"]
style_header_row(ws_detail)

header = [cell.value for cell in ws_detail[1]]
priority_col = header.index("Prioridad") + 1

for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row):
    priority_val = str(row[priority_col - 1].value or "")
    if "P1" in priority_val:
        bg = COLOR_P1
    elif "P2" in priority_val:
        bg = COLOR_P2
    elif "P3" in priority_val:
        bg = COLOR_P3
    elif "P4" in priority_val:
        bg = COLOR_P4
    else:
        bg = None

    for cell in row:
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        elif cell.row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=COLOR_GRAY)
        cell.alignment = Alignment(horizontal="center")
        add_border(cell)

ws_detail.freeze_panes = "A2"
auto_width(ws_detail)

# ── Formato pestaña RESUMEN ───────────────────────────────────────────────
ws_summary = wb["RESUMEN"]
style_header_row(ws_summary)

ws_summary.cell(row=13, column=1).value = "TOP 10 ROLES MAS CRITICOS"
ws_summary.cell(row=13, column=1).font = Font(bold=True, size=12, color=COLOR_HEADER)

style_header_row(ws_summary, row=15)
auto_width(ws_summary)

for row in ws_summary.iter_rows(min_row=2, max_row=12, min_col=1, max_col=1):
    for cell in row:
        cell.font = Font(bold=True)

# ── Formato pestaña SUPUESTOS ─────────────────────────────────────────────
ws_supuestos = wb["SUPUESTOS"]
style_header_row(ws_supuestos)

for row in ws_supuestos.iter_rows(min_row=2, max_row=ws_supuestos.max_row):
    val = str(row[0].value or "")
    if val.isupper() and val != "":
        for cell in row:
            cell.font = Font(bold=True, color=COLOR_HEADER, size=11)
            cell.fill = PatternFill("solid", fgColor="DCE6F1")

auto_width(ws_supuestos)

wb.save(OUTPUT)
print("Formato aplicado correctamente.")
print(f"\nArchivo final: {OUTPUT}")
print("Pestanas: DETALLE · RESUMEN · SUPUESTOS")
